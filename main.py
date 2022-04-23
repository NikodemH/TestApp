import TestCaseDB
import TestCaseDefinitionDB
import TestPlanDB
from openpyxl import load_workbook
import argparse

parser = argparse.ArgumentParser()
group = parser.add_mutually_exclusive_group()
group.add_argument('-r', action='store_true')
group.add_argument('-w', action='store_true')
parser.add_argument('-a', action="store")
parser.add_argument('-b', action="store")
parser.add_argument('-c', action="store")
args = parser.parse_args()


if args.r:
    if args.a == 'TP':
        if args.b == 'TC':
            wb = load_workbook('TestPlan.xlsx')
            pos = (TestPlanDB.db.find(args.c))
            cel = wb.active.cell(pos, 2)
            TestPlanDB.db.read_multiple(cel)
            # print(TestPlanDB.db.read_multiple(wb.active.cell(pos, 2)))
if args.w:
    print("NO")
# TestCaseDB.db.read_multiple("A1")
# TestCaseDefinitionDB.db.read_multiple("A1")
