from openpyxl import Workbook


wb = Workbook()
ws = wb.active

# Read and split cell content separated by spaces
def read_multiple(cell):
    text = ws[cell]
    text_split = text.value.split()
    print(text_split)

# Find first empty spot to write in
def find_empty():
    i = 2
    c = ws.cell(i, 1)
    while c.value is not None:
        i += 1
        c = ws.cell(i, 1)
    return i

# Find cell with specific value in first column (ID)
def find(inumber):
    i = 2
    c = ws.cell(i, 1)
    while c.value != inumber:
        i += 1
        c = ws.cell(i, 1)
        if c.value is None:
            return None
    return i


